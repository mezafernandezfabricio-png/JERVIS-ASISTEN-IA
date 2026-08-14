# -*- coding: utf-8 -*-
"""
excel_tools.py — Gestor y creador profesional de hojas de cálculo Excel (.xlsx) y CSV para JARVIS.
Genera tablas completas, formateadas y con datos realistas.
Guarda SIEMPRE en el Escritorio del usuario con apertura automática.
"""

import os
import sys
import re
import csv
import json
from pathlib import Path
from datetime import datetime

BASE_DIR = Path(__file__).resolve().parent.parent
API_FILE = BASE_DIR / "config" / "api_keys.json"

def _get_desktop_dir() -> Path:
    desktop = Path.home() / "Desktop"
    if not desktop.exists():
        onedrive_desktop = Path.home() / "OneDrive" / "Desktop"
        if onedrive_desktop.exists():
            return onedrive_desktop
        desktop.mkdir(parents=True, exist_ok=True)
    return desktop

def _get_api_key() -> str:
    try:
        if API_FILE.exists():
            return json.loads(API_FILE.read_text(encoding="utf-8")).get("gemini_api_key", "")
    except Exception:
        pass
    return ""

def _sanitize_filename(name: str) -> str:
    cleaned = re.sub(r'[\\/*?:"<>|]', "", name).strip()
    return cleaned if cleaned else f"Planilla_Excel_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

def _generate_table_data_if_needed(title: str, headers: list, data: list) -> tuple:
    """Si no se proporcionan datos suficientes, genera una tabla estructurada y realista con IA."""
    if data and len(data) >= 3 and headers:
        return headers, data

    api_key = _get_api_key()
    if not api_key:
        return headers, data

    try:
        from google import genai
        client = genai.Client(api_key=api_key)
        prompt = (
            f"Actúa como un experto analista de datos de JARVIS. Genera una tabla de datos completa, realista y detallada para una hoja de cálculo Excel.\n"
            f"Tema / Título: '{title}'\n"
            f"Encabezados sugeridos: {headers}\n"
            f"Datos sugeridos: {data}\n\n"
            "REQUISITOS OBLIGATORIOS:\n"
            "Devuelve ÚNICAMENTE un objeto JSON válido con este esquema exacto:\n"
            "{\n"
            "  \"headers\": [\"Columna 1\", \"Columna 2\", \"Columna 3\", \"Columna 4\", \"Columna 5\"],\n"
            "  \"rows\": [\n"
            "    [\"Valor 1\", 100, \"Texto\", \"Otro dato\", 50.5],\n"
            "    [\"Valor 2\", 200, \"Texto\", \"Otro dato\", 75.0]\n"
            "  ]\n"
            "}\n"
            "Genera entre 8 y 15 filas con información variada, coherente y precisa en español."
        )
        resp = client.models.generate_content(model="gemini-2.5-flash", contents=prompt)
        text = resp.text.strip()
        m = re.search(r'\{[\s\S]*\}', text)
        if m:
            parsed = json.loads(m.group(0))
            new_headers = parsed.get("headers", headers)
            new_rows = parsed.get("rows", data)
            if new_headers and new_rows:
                return new_headers, new_rows
    except Exception:
        pass
    return headers, data

def excel_tools(parameters: dict = None, player=None, speak=None, **kwargs) -> str:
    """
    Crea, formatea y organiza hojas de cálculo Excel (.xlsx) o CSV en el Escritorio.
    Parámetros:
        - action: 'create' (default), 'read'
        - filename / title / nombre: Nombre del archivo Excel
        - headers: Lista de encabezados de columnas
        - data / rows / rows_data / content: Lista de filas o datos
        - sheet_name: Nombre de la pestaña (opcional)
        - open_file / abrir: Si debe abrir el archivo (default True)
    """
    params = parameters or {}
    action = params.get("action", "create").lower()
    title = params.get("title") or params.get("filename") or params.get("nombre") or "Planilla_JARVIS"
    headers = params.get("headers") or []
    data = params.get("data") or params.get("rows") or params.get("rows_data") or params.get("content") or []
    sheet_name = params.get("sheet_name", "Datos")
    open_file = params.get("open_file", params.get("abrir", True))
    
    clean_title = _sanitize_filename(title)
    if clean_title.lower().endswith(".xlsx"):
        clean_title = clean_title[:-5]
    elif clean_title.lower().endswith(".csv"):
        clean_title = clean_title[:-4]

    desktop = _get_desktop_dir()
    excel_path = desktop / f"{clean_title}.xlsx"

    if action == "read":
        return _read_excel_or_csv(excel_path if excel_path.exists() else desktop / f"{clean_title}.csv")

    if player:
        try: player.write_log(f"📊 Generando libro de cálculo Excel: '{clean_title}.xlsx'...")
        except: pass

    # Parsear datos de entrada si vienen como texto
    if isinstance(data, str):
        parsed_rows = []
        for line in data.strip().split("\n"):
            line_str = line.strip()
            if not line_str: continue
            if "\t" in line_str:
                parsed_rows.append([cell.strip() for cell in line_str.split("\t")])
            elif "," in line_str:
                parsed_rows.append([cell.strip() for cell in line_str.split(",")])
            elif ";" in line_str:
                parsed_rows.append([cell.strip() for cell in line_str.split(";")])
            elif "|" in line_str:
                parsed_rows.append([cell.strip() for cell in line_str.split("|") if cell.strip()])
            else:
                parsed_rows.append([line_str])
        data = parsed_rows

    if isinstance(headers, str):
        headers = [h.strip() for h in re.split(r'[,;|\t]', headers) if h.strip()]

    # Generar tabla rica con IA si la información es escasa
    headers, data = _generate_table_data_if_needed(title, headers, data)

    # Si aún no hay encabezados pero sí filas
    if not headers and data and isinstance(data, list) and len(data) > 1 and isinstance(data[0], list):
        headers = data[0]
        data = data[1:]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:30]

        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Segoe UI", size=10.5, color="1E293B")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        row_idx = 1
        if headers:
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(h))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            ws.row_dimensions[row_idx].height = 26
            row_idx += 1

        for r_idx, row in enumerate(data, row_idx):
            is_alt = (r_idx % 2 == 0)
            if isinstance(row, list):
                for col_idx, val in enumerate(row, 1):
                    clean_val = val
                    if isinstance(val, str):
                        val_s = val.strip()
                        if re.match(r'^-?\d+(\.\d+)?$', val_s):
                            try: clean_val = float(val_s) if '.' in val_s else int(val_s)
                            except: clean_val = val
                    cell = ws.cell(row=r_idx, column=col_idx, value=clean_val)
                    cell.font = regular_font
                    cell.border = thin_border
                    if is_alt: cell.fill = alt_fill
                    cell.alignment = Alignment(vertical="center", horizontal="right" if isinstance(clean_val, (int, float)) else "left")
            else:
                cell = ws.cell(row=r_idx, column=1, value=str(row))
                cell.font = regular_font
                cell.border = thin_border
            ws.row_dimensions[r_idx].height = 20

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        wb.save(str(excel_path))

    except Exception:
        csv_path = desktop / f"{clean_title}.csv"
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            if headers: writer.writerow(headers)
            for row in data:
                writer.writerow(row if isinstance(row, list) else [row])
        excel_path = csv_path

    if player:
        try: player.write_log(f"📊 Hoja de cálculo creada en Escritorio: {excel_path.name}")
        except: pass

    if open_file and excel_path.exists():
        try: os.startfile(str(excel_path))
        except: pass

    return f"Hoja de cálculo creada exitosamente en tu Escritorio: '{excel_path.name}' con toda la información solicitada."

def _read_excel_or_csv(file_path: Path) -> str:
    if not file_path.exists():
        return f"No se encontró el archivo: {file_path}"
    try:
        import pandas as pd
        df = pd.read_excel(str(file_path)) if file_path.suffix.lower() == ".xlsx" else pd.read_csv(str(file_path))
        return f"Contenido de '{file_path.name}':\n{df.head(20).to_string()}"
    except:
        return f"Archivo localizado en '{file_path}'."
